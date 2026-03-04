"""XlsxExtractor — extracts text from XLSX files using openpyxl."""

from __future__ import annotations

from typing import Any, ClassVar

from ..core.base import BaseExtractor, ExtractionError, ExtractionResult


class XlsxExtractor(BaseExtractor):
    """Extracts text from Excel XLSX files using openpyxl."""

    supported_extensions: ClassVar[set[str]] = {".xlsx"}
    required_packages: ClassVar[set[str]] = {"openpyxl"}

    def extract(self, source: str) -> ExtractionResult:
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise ExtractionError(
                "openpyxl is required for XLSX extraction: pip install openpyxl",
                source=source,
                cause=e,
            )

        try:
            wb = load_workbook(source, read_only=True, data_only=True)
        except Exception as e:
            raise ExtractionError(
                f"Failed to open XLSX {source}: {e}", source=source, cause=e
            )

        text_parts: list[str] = []
        md_parts: list[str] = []
        total_rows = 0
        metadata: dict[str, Any] = {}

        try:
            metadata["Sheets"] = ", ".join(wb.sheetnames)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                text_parts.append(f"[{sheet_name}]")
                md_parts.append(f"## Sheet: {sheet_name}")

                rows: list[list[str]] = []
                for row in ws.iter_rows():
                    values = [self._cell_value(cell) for cell in row]
                    # Skip completely empty rows
                    if not any(v for v in values):
                        continue
                    total_rows += 1
                    text_parts.append("\t".join(values))
                    rows.append(values)

                if rows:
                    md_parts.append(self._rows_to_markdown(rows))
        finally:
            wb.close()

        metadata["Total rows"] = total_rows
        full_text = "\n".join(text_parts)
        markdown_text = "\n\n".join(md_parts)

        return ExtractionResult(
            text=full_text,
            source=source,
            source_type="xlsx",
            extractor_name=self.__class__.__name__,
            metadata=metadata,
            markdown_text=markdown_text or None,
        )

    @staticmethod
    def _cell_value(cell: Any) -> str:
        """Extract string value from a cell, handling None."""
        val = cell.value
        if val is None:
            return ""
        return str(val).strip()

    @staticmethod
    def _rows_to_markdown(rows: list[list[str]]) -> str:
        """Render rows as a markdown table. First row is treated as the header."""
        if not rows:
            return ""

        cols = max(len(row) for row in rows)
        # Pad rows to same width
        padded = [row + [""] * (cols - len(row)) for row in rows]

        header = padded[0]
        lines = [
            "| " + " | ".join(header) + " |",
            "| " + " | ".join("---" for _ in range(cols)) + " |",
        ]
        for row in padded[1:]:
            lines.append("| " + " | ".join(row) + " |")
        return "\n".join(lines)
